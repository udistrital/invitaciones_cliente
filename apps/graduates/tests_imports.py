import io

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.ceremonies.models import Ceremony
from apps.graduates.imports import (
    TEMPLATE_HEADERS,
    build_graduate_template_workbook,
    confirm_graduate_import_batch,
    create_graduate_import_batch,
)
from apps.graduates.models import Graduate, GraduateImportBatch
from apps.invitations.models import Invitation
from apps.invitations.services import issue_invitations_for_graduate


class GraduateImportServiceTest(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="importador",
            password="secret123",
            is_staff=True,
        )
        self.ceremony = Ceremony.objects.create(
            code="GRADOS-2026-01",
            name="Ceremonia Central",
            scheduled_at=timezone.now(),
            venue="Auditorio Sabio Caldas",
        )

    def build_excel_file(self, rows, headers=TEMPLATE_HEADERS, name="graduandos.xlsx"):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "graduandos"
        worksheet.append(list(headers))
        for row in rows:
            worksheet.append(list(row))

        buffer = io.BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            name,
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def test_template_workbook_contains_expected_headers(self):
        workbook = load_workbook(io.BytesIO(build_graduate_template_workbook()))
        worksheet = workbook.active

        self.assertEqual(worksheet.title, "graduandos")
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            list(TEMPLATE_HEADERS),
        )

    def test_batch_preview_reports_missing_required_columns(self):
        upload = self.build_excel_file(
            rows=[("CC", "123", "Laura Perez")],
            headers=("tipo_documento", "numero_documento", "nombre_completo"),
        )

        batch = create_graduate_import_batch(
            ceremony=self.ceremony,
            uploaded_file=upload,
            uploaded_by=self.user,
        )

        self.assertEqual(batch.status, GraduateImportBatch.Status.VALIDATED)
        self.assertFalse(batch.preview_payload["can_confirm"])
        self.assertTrue(batch.preview_payload["file_errors"])

    def test_batch_preview_detects_conflict_between_document_and_student_code(self):
        Graduate.objects.create(
            ceremony=self.ceremony,
            student_code="20260001",
            document_number="10000001",
            full_name="Laura Perez",
            academic_program="Ingenieria de Sistemas",
            invitation_quota=3,
        )
        Graduate.objects.create(
            ceremony=self.ceremony,
            student_code="20260002",
            document_number="10000002",
            full_name="Carlos Ruiz",
            academic_program="Ingenieria Industrial",
            invitation_quota=3,
        )
        upload = self.build_excel_file(
            rows=[
                (
                    "20260002",
                    "CC",
                    "10000001",
                    "Laura Perez Actualizada",
                    "laura@example.com",
                    "Ingenieria de Sistemas",
                )
            ]
        )

        batch = create_graduate_import_batch(
            ceremony=self.ceremony,
            uploaded_file=upload,
            uploaded_by=self.user,
        )

        self.assertEqual(batch.rows_error, 1)
        self.assertEqual(batch.preview_payload["rows"][0]["action"], "error")

    def test_confirm_import_creates_graduates_and_three_invitations_each(self):
        upload = self.build_excel_file(
            rows=[
                (
                    "20260001",
                    "CC",
                    "10000001",
                    "Laura Perez",
                    "laura@example.com",
                    "Ingenieria de Sistemas",
                ),
                (
                    "20260002",
                    "CC",
                    "10000002",
                    "Carlos Ruiz",
                    "carlos@example.com",
                    "Ingenieria Industrial",
                ),
            ]
        )
        batch = create_graduate_import_batch(
            ceremony=self.ceremony,
            uploaded_file=upload,
            uploaded_by=self.user,
        )

        confirm_graduate_import_batch(batch=batch, confirmed_by=self.user)

        self.assertEqual(batch.status, GraduateImportBatch.Status.CONFIRMED)
        self.assertEqual(Graduate.objects.filter(ceremony=self.ceremony).count(), 2)
        self.assertEqual(
            Invitation.objects.filter(graduate__ceremony=self.ceremony).count(),
            6,
        )
        self.assertEqual(batch.graduates_created, 2)
        self.assertEqual(batch.invitations_created, 6)

    def test_confirm_import_updates_existing_graduate_and_completes_missing_invitations(self):
        graduate = Graduate.objects.create(
            ceremony=self.ceremony,
            student_code="20260001",
            document_type="CC",
            document_number="10000001",
            full_name="Laura Perez",
            academic_program="Ingenieria de Sistemas",
            invitation_quota=2,
        )
        issue_invitations_for_graduate(graduate)
        graduate.email = ""
        graduate.save(update_fields=["email"])

        upload = self.build_excel_file(
            rows=[
                (
                    "20260001",
                    "CC",
                    "10000001",
                    "Laura Perez",
                    "laura@example.com",
                    "Ingenieria de Sistemas",
                )
            ]
        )
        batch = create_graduate_import_batch(
            ceremony=self.ceremony,
            uploaded_file=upload,
            uploaded_by=self.user,
        )

        self.assertEqual(
            batch.preview_payload["rows"][0]["action"],
            "actualizar_y_completar",
        )

        confirm_graduate_import_batch(batch=batch, confirmed_by=self.user)

        graduate.refresh_from_db()
        self.assertEqual(graduate.invitation_quota, 3)
        self.assertEqual(graduate.email, "laura@example.com")
        self.assertEqual(graduate.invitations.count(), 3)
        self.assertEqual(batch.graduates_updated, 1)
        self.assertEqual(batch.invitations_created, 1)

    def test_preview_rejects_existing_graduate_with_more_than_three_invitations(self):
        graduate = Graduate.objects.create(
            ceremony=self.ceremony,
            student_code="20260001",
            document_type="CC",
            document_number="10000001",
            full_name="Laura Perez",
            academic_program="Ingenieria de Sistemas",
            invitation_quota=4,
        )
        for sequence_number in range(1, 5):
            Invitation.objects.create(
                graduate=graduate,
                sequence_number=sequence_number,
            )

        upload = self.build_excel_file(
            rows=[
                (
                    "20260001",
                    "CC",
                    "10000001",
                    "Laura Perez",
                    "laura@example.com",
                    "Ingenieria de Sistemas",
                )
            ]
        )

        batch = create_graduate_import_batch(
            ceremony=self.ceremony,
            uploaded_file=upload,
            uploaded_by=self.user,
        )

        self.assertFalse(batch.preview_payload["can_confirm"])
        self.assertIn(
            "mas de 3 invitaciones",
            batch.preview_payload["rows"][0]["errors"][0],
        )
