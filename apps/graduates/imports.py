import hashlib
import io
import logging
from collections import Counter

from django.core.exceptions import ValidationError
from django.core.validators import EmailValidator
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from apps.graduates.models import Graduate, GraduateImportBatch
from apps.invitations.services import issue_invitations_for_graduate

logger = logging.getLogger(__name__)

TEMPLATE_HEADERS = (
    "codigo_estudiantil",
    "tipo_documento",
    "numero_documento",
    "nombre_completo",
    "correo_institucional",
    "programa_academico",
)

REQUIRED_HEADERS = (
    "codigo_estudiantil",
    "tipo_documento",
    "numero_documento",
    "nombre_completo",
    "programa_academico",
)

EMAIL_VALIDATOR = EmailValidator()


def build_graduate_template_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "graduandos"
    worksheet.append(list(TEMPLATE_HEADERS))

    for index, header in enumerate(TEMPLATE_HEADERS, start=1):
        worksheet.cell(row=1, column=index).value = header
        worksheet.cell(row=1, column=index).font = worksheet.cell(
            row=1,
            column=index,
        ).font.copy(bold=True)
        worksheet.column_dimensions[get_column_letter(index)].width = 24

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_graduate_import_batch(*, ceremony, uploaded_file, uploaded_by=None):
    source = read_graduate_import_source(uploaded_file)
    preview = build_graduate_import_preview(
        ceremony=ceremony,
        source_rows=source["source_rows"],
        sheet_name=source["sheet_name"],
        file_errors=source["file_errors"],
    )

    batch = GraduateImportBatch.objects.create(
        ceremony=ceremony,
        uploaded_by=uploaded_by,
        source_filename=source["source_filename"],
        file_sha256=source["file_sha256"],
        rows_total=preview["rows_total"],
        rows_valid=preview["rows_valid"],
        rows_error=preview["rows_error"],
        preview_payload=preview,
    )
    logger.info(
        "Graduate import batch validated batch_id=%s ceremony_id=%s ceremony_code=%s filename=%s rows_total=%s rows_valid=%s rows_error=%s uploaded_by_id=%s",
        batch.pk,
        ceremony.pk,
        ceremony.code,
        batch.source_filename,
        batch.rows_total,
        batch.rows_valid,
        batch.rows_error,
        getattr(uploaded_by, "pk", None),
    )
    return batch


def confirm_graduate_import_batch(*, batch, confirmed_by=None):
    preview = build_graduate_import_preview(
        ceremony=batch.ceremony,
        source_rows=batch.preview_payload.get("source_rows", []),
        sheet_name=batch.preview_payload.get("sheet_name", ""),
        file_errors=batch.preview_payload.get("file_errors", []),
    )

    if not preview["can_confirm"]:
        raise ValidationError(
            "La importacion no se puede confirmar porque el lote tiene errores."
        )

    try:
        with transaction.atomic():
            graduates_created = 0
            graduates_updated = 0
            invitations_created = 0

            for row in preview["rows"]:
                graduate, was_created, was_updated = apply_graduate_import_row(
                    ceremony=batch.ceremony,
                    row=row,
                )
                graduates_created += int(was_created)
                graduates_updated += int(was_updated)

                invitations_before = graduate.invitations.count()
                issue_invitations_for_graduate(graduate)
                invitations_after = graduate.invitations.count()
                invitations_created += invitations_after - invitations_before

            batch.status = GraduateImportBatch.Status.CONFIRMED
            batch.confirmed_by = confirmed_by
            batch.confirmed_at = timezone.now()
            batch.rows_total = preview["rows_total"]
            batch.rows_valid = preview["rows_valid"]
            batch.rows_error = preview["rows_error"]
            batch.graduates_created = graduates_created
            batch.graduates_updated = graduates_updated
            batch.invitations_created = invitations_created
            batch.preview_payload = preview
            batch.failure_message = ""
            batch.save()
            logger.info(
                "Graduate import batch confirmed batch_id=%s ceremony_id=%s ceremony_code=%s graduates_created=%s graduates_updated=%s invitations_created=%s confirmed_by_id=%s",
                batch.pk,
                batch.ceremony_id,
                batch.ceremony.code,
                batch.graduates_created,
                batch.graduates_updated,
                batch.invitations_created,
                getattr(confirmed_by, "pk", None),
            )
    except Exception as exc:
        GraduateImportBatch.objects.filter(pk=batch.pk).update(
            status=GraduateImportBatch.Status.FAILED,
            failure_message=str(exc),
            preview_payload=preview,
            rows_total=preview["rows_total"],
            rows_valid=preview["rows_valid"],
            rows_error=preview["rows_error"],
        )
        logger.exception(
            "Graduate import batch failed batch_id=%s ceremony_id=%s ceremony_code=%s",
            batch.pk,
            batch.ceremony_id,
            batch.ceremony.code,
        )
        raise

    return batch


def read_graduate_import_source(uploaded_file):
    source_filename = getattr(uploaded_file, "name", "") or "graduandos.xlsx"
    file_bytes = uploaded_file.read()
    file_errors = []

    if not source_filename.lower().endswith(".xlsx"):
        file_errors.append("El archivo debe tener extension .xlsx.")

    if not file_bytes:
        file_errors.append("El archivo esta vacio.")

    source_rows = []
    sheet_name = ""

    if not file_errors:
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception:
            workbook = None
            file_errors.append("No fue posible leer el archivo Excel.")

        if workbook is not None:
            worksheet, missing_headers = find_first_valid_worksheet(workbook)
            if worksheet is None:
                file_errors.append(
                    "No se encontro una hoja con las columnas minimas requeridas: "
                    + ", ".join(REQUIRED_HEADERS)
                    + "."
                )
            else:
                sheet_name = worksheet.title
                if missing_headers:
                    file_errors.append(
                        "Faltan columnas obligatorias: " + ", ".join(missing_headers) + "."
                    )
                else:
                    source_rows = extract_source_rows(worksheet)

    return {
        "source_filename": source_filename,
        "file_sha256": hashlib.sha256(file_bytes).hexdigest(),
        "file_errors": file_errors,
        "sheet_name": sheet_name,
        "source_rows": source_rows,
    }


def find_first_valid_worksheet(workbook):
    fallback_missing = tuple(REQUIRED_HEADERS)

    for worksheet in workbook.worksheets:
        header_map = build_header_map(worksheet)
        missing_headers = tuple(
            header for header in REQUIRED_HEADERS if header not in header_map
        )
        if not missing_headers:
            return worksheet, ()
        fallback_missing = missing_headers

    return None, fallback_missing


def build_header_map(worksheet):
    header_map = {}
    for index, cell in enumerate(worksheet[1], start=1):
        header_name = normalize_header(cell.value)
        if header_name:
            header_map[header_name] = index
    return header_map


def extract_source_rows(worksheet):
    header_map = build_header_map(worksheet)
    source_rows = []

    for row_index in range(2, worksheet.max_row + 1):
        row_data = {
            "row_number": row_index,
            "student_code": normalize_value(
                worksheet.cell(
                    row=row_index,
                    column=header_map["codigo_estudiantil"],
                ).value
            ),
            "document_type": normalize_value(
                worksheet.cell(
                    row=row_index,
                    column=header_map["tipo_documento"],
                ).value
            ),
            "document_number": normalize_value(
                worksheet.cell(
                    row=row_index,
                    column=header_map["numero_documento"],
                ).value
            ),
            "full_name": normalize_value(
                worksheet.cell(
                    row=row_index,
                    column=header_map["nombre_completo"],
                ).value
            ),
            "email": normalize_value(
                worksheet.cell(
                    row=row_index,
                    column=header_map.get("correo_institucional", 0),
                ).value
                if header_map.get("correo_institucional")
                else ""
            ),
            "academic_program": normalize_value(
                worksheet.cell(
                    row=row_index,
                    column=header_map["programa_academico"],
                ).value
            ),
        }

        if any(
            row_data[field]
            for field in (
                "student_code",
                "document_type",
                "document_number",
                "full_name",
                "email",
                "academic_program",
            )
        ):
            source_rows.append(row_data)

    return source_rows


def build_graduate_import_preview(*, ceremony, source_rows, sheet_name="", file_errors=None):
    file_errors = list(file_errors or [])
    existing_graduates = Graduate.objects.filter(ceremony=ceremony).prefetch_related(
        "invitations"
    )
    existing_by_document = {
        graduate.document_number: graduate for graduate in existing_graduates
    }
    existing_by_student_code = {
        graduate.student_code: graduate
        for graduate in existing_graduates
        if graduate.student_code
    }

    document_counts = Counter(
        row["document_number"] for row in source_rows if row["document_number"]
    )
    student_code_counts = Counter(
        row["student_code"] for row in source_rows if row["student_code"]
    )

    rows = []
    rows_valid = 0
    rows_error = 0

    for source_row in source_rows:
        row_preview = build_row_preview(
            source_row=source_row,
            existing_by_document=existing_by_document,
            existing_by_student_code=existing_by_student_code,
            document_counts=document_counts,
            student_code_counts=student_code_counts,
        )
        rows.append(row_preview)
        if row_preview["errors"]:
            rows_error += 1
        else:
            rows_valid += 1

    return {
        "sheet_name": sheet_name,
        "file_errors": file_errors,
        "source_rows": source_rows,
        "rows": rows,
        "rows_total": len(source_rows),
        "rows_valid": rows_valid,
        "rows_error": rows_error,
        "can_confirm": bool(source_rows) and not file_errors and rows_error == 0,
    }


def build_row_preview(
    *,
    source_row,
    existing_by_document,
    existing_by_student_code,
    document_counts,
    student_code_counts,
):
    errors = []
    row_data = {
        "student_code": source_row["student_code"],
        "document_type": source_row["document_type"],
        "document_number": source_row["document_number"],
        "full_name": source_row["full_name"],
        "email": source_row["email"],
        "academic_program": source_row["academic_program"],
    }

    for field_name, label in (
        ("student_code", "codigo_estudiantil"),
        ("document_type", "tipo_documento"),
        ("document_number", "numero_documento"),
        ("full_name", "nombre_completo"),
        ("academic_program", "programa_academico"),
    ):
        if not row_data[field_name]:
            errors.append(f"El campo {label} es obligatorio.")

    if row_data["email"]:
        try:
            EMAIL_VALIDATOR(row_data["email"])
        except ValidationError:
            errors.append("El correo_institucional no tiene un formato valido.")

    if row_data["document_number"] and document_counts[row_data["document_number"]] > 1:
        errors.append("El numero_documento esta duplicado dentro del archivo.")

    if row_data["student_code"] and student_code_counts[row_data["student_code"]] > 1:
        errors.append("El codigo_estudiantil esta duplicado dentro del archivo.")

    existing_by_document_number = existing_by_document.get(row_data["document_number"])
    existing_by_student_identifier = existing_by_student_code.get(row_data["student_code"])

    if (
        existing_by_document_number
        and existing_by_student_identifier
        and existing_by_document_number.pk != existing_by_student_identifier.pk
    ):
        errors.append(
            "El numero_documento y el codigo_estudiantil apuntan a graduandos distintos."
        )

    existing_graduate = existing_by_document_number or existing_by_student_identifier
    existing_invitation_count = existing_graduate.invitations.count() if existing_graduate else 0

    if existing_graduate and existing_invitation_count > 3:
        errors.append(
            "El graduando ya tiene mas de 3 invitaciones y requiere revision manual."
        )

    changed_fields = []
    action = "crear"
    invitations_to_create = 3

    if existing_graduate:
        for source_field, model_field in (
            ("student_code", "student_code"),
            ("document_type", "document_type"),
            ("document_number", "document_number"),
            ("full_name", "full_name"),
            ("email", "email"),
            ("academic_program", "academic_program"),
        ):
            if getattr(existing_graduate, model_field) != row_data[source_field]:
                changed_fields.append(source_field)

        needs_update = bool(changed_fields or existing_graduate.invitation_quota != 3)
        invitations_to_create = max(0, 3 - existing_invitation_count)

        if invitations_to_create > 0:
            action = "actualizar_y_completar"
        elif needs_update:
            action = "actualizar"
        else:
            action = "sin_cambios"

    if errors:
        action = "error"
        invitations_to_create = 0

    return {
        "row_number": source_row["row_number"],
        "data": row_data,
        "errors": errors,
        "action": action,
        "changed_fields": changed_fields,
        "existing_graduate_id": existing_graduate.pk if existing_graduate else None,
        "existing_invitation_count": existing_invitation_count,
        "invitations_to_create": invitations_to_create,
    }


def apply_graduate_import_row(*, ceremony, row):
    if row["errors"]:
        raise ValidationError("No es posible confirmar un lote con filas invalidas.")

    graduate = resolve_existing_graduate(
        ceremony=ceremony,
        row_data=row["data"],
        existing_graduate_id=row.get("existing_graduate_id"),
    )

    if graduate is None:
        graduate = Graduate.objects.create(
            ceremony=ceremony,
            student_code=row["data"]["student_code"],
            document_type=row["data"]["document_type"],
            document_number=row["data"]["document_number"],
            full_name=row["data"]["full_name"],
            academic_program=row["data"]["academic_program"],
            email=row["data"]["email"],
            invitation_quota=3,
        )
        return graduate, True, False

    if graduate.invitations.count() > 3:
        raise ValidationError(
            "El graduando ya tiene mas de 3 invitaciones y no puede importarse."
        )

    changed = False
    for source_field, model_field in (
        ("student_code", "student_code"),
        ("document_type", "document_type"),
        ("document_number", "document_number"),
        ("full_name", "full_name"),
        ("email", "email"),
        ("academic_program", "academic_program"),
    ):
        new_value = row["data"][source_field]
        if getattr(graduate, model_field) != new_value:
            setattr(graduate, model_field, new_value)
            changed = True

    if graduate.invitation_quota != 3:
        graduate.invitation_quota = 3
        changed = True

    if changed:
        graduate.save()

    return graduate, False, changed


def resolve_existing_graduate(*, ceremony, row_data, existing_graduate_id=None):
    graduate_by_id = None
    if existing_graduate_id:
        graduate_by_id = Graduate.objects.filter(
            pk=existing_graduate_id,
            ceremony=ceremony,
        ).first()

    graduate_by_document = None
    if row_data["document_number"]:
        graduate_by_document = Graduate.objects.filter(
            ceremony=ceremony,
            document_number=row_data["document_number"],
        ).first()

    graduate_by_student_code = None
    if row_data["student_code"]:
        graduate_by_student_code = Graduate.objects.filter(
            ceremony=ceremony,
            student_code=row_data["student_code"],
        ).first()

    matches = [
        graduate
        for graduate in (
            graduate_by_id,
            graduate_by_document,
            graduate_by_student_code,
        )
        if graduate is not None
    ]
    unique_ids = {graduate.pk for graduate in matches}

    if len(unique_ids) > 1:
        raise ValidationError(
            "La fila no se puede confirmar porque documento y codigo apuntan a registros distintos."
        )

    if matches:
        return matches[0]
    return None


def normalize_header(value):
    return normalize_value(value).lower()


def normalize_value(value):
    if value is None:
        return ""
    return str(value).strip()
